"""Carga masiva desde Excel: subir+previsualizar, confirmar mapeo, aprobar.

No reutiliza el worker de `bulk_import.py` (ese es para sincronizar
respuestas de formulario ya estructuradas via API/dispositivo). Este motor
importa filas crudas de un archivo .xlsx hacia participantes, usuarios,
asignaciones usuario-proyecto-rol (ver docs/103) o registros historicos de
un formulario (ver docs/104), validando y dejando reporte de errores antes
de decidir importar.
"""

import json
from datetime import datetime
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.time import utc_now
from app.models.builder import BuilderComponent, BuilderTemplate
from app.models.excel_import import ExcelImportJob
from app.models.identity import Role
from app.models.runtime_record import RuntimeRecord
from app.schemas.assignment import AssignmentCreate
from app.schemas.excel_import import ExcelImportJobRead, ExcelImportTargetField
from app.schemas.identity import UserCreate
from app.schemas.participants import ParticipantCreate
from app.schemas.runtime_record import RuntimeRecordCreate, RuntimeValueCreate
from app.services.assignment_service import assignment_service
from app.services.identity_service import identity_service
from app.services.participant_service import participant_service
from app.services.runtime_record_service import runtime_record_service

PREVIEW_ROW_LIMIT = 20

# entity_type="records": solo campos escalares simples (una celda de Excel =
# un valor). Se excluyen archivos/firma/GPS/subformularios/multi-seleccion/
# campos calculados o de sistema -- documentado como pendiente en docs/104.
RECORD_SIMPLE_FIELD_TYPES = {
    "TEXT", "TEXTAREA", "DOCUMENT_ID", "EMAIL", "PHONE", "URL",
    "SELECT", "DROPDOWN", "DATE", "TIME", "DATETIME", "YEAR", "MONTH", "WEEK",
    "NUMBER", "INTEGER", "DECIMAL", "PERCENTAGE", "CURRENCY",
    "RATING", "NPS", "RANGE", "BOOLEAN",
}
RECORD_NUMERIC_FIELD_TYPES = {"NUMBER", "INTEGER", "DECIMAL", "PERCENTAGE", "CURRENCY", "RATING", "NPS", "RANGE"}
RECORD_BOOLEAN_TRUE_VALUES = {"true", "verdadero", "si", "sí", "1", "x", "yes"}
RECORD_BOOLEAN_FALSE_VALUES = {"false", "falso", "no", "0"}
RECORD_VALID_STATUSES = {"draft", "submitted", "approved", "rejected", "archived"}

# Campos destino reservados para entity_type="records" -- nombres poco
# probables de colisionar con un BuilderComponent.name real.
META_STATUS_FIELD = "_meta_status"
META_CREATED_AT_FIELD = "_meta_created_at"

ENTITY_ALIASES: dict[str, dict[str, str]] = {
    "participants": {
        "documento": "document_id",
        "cedula": "document_id",
        "identificacion": "document_id",
        "nombre": "full_name",
        "nombre completo": "full_name",
        "codigo": "external_code",
        "codigo externo": "external_code",
        "tipo": "participant_type",
    },
    "users": {
        "documento": "document_id",
        "cedula": "document_id",
        "nombre": "full_name",
        "nombre completo": "full_name",
        "correo": "email",
        "email": "email",
        "correo electronico": "email",
        "telefono": "phone",
        "celular": "phone",
    },
    "assignments": {
        "correo": "email",
        "email": "email",
        "correo electronico": "email",
        "usuario": "email",
        "rol": "role_name",
        "nombre del rol": "role_name",
        "role": "role_name",
        "estado": "status",
        "status": "status",
    },
}

PARTICIPANT_TARGET_FIELDS = {"document_id", "full_name", "external_code", "participant_type"}
USER_TARGET_FIELDS = {"document_id", "full_name", "email", "phone"}
ASSIGNMENT_TARGET_FIELDS = {"email", "role_name", "status"}

REQUIRED_FIELDS = {
    "participants": {"full_name"},
    "users": {"full_name", "document_id", "email"},
    "assignments": {"email", "role_name"},
}


def _template_simple_components(db: Session, template_id: str) -> list[BuilderComponent]:
    return db.query(BuilderComponent).filter(
        BuilderComponent.template_id == template_id,
        BuilderComponent.component_type.in_(RECORD_SIMPLE_FIELD_TYPES),
    ).order_by(BuilderComponent.sort_order).all()


def _record_target_fields(db: Session, template_id: str) -> list[ExcelImportTargetField]:
    fields = [ExcelImportTargetField(name=c.name, label=c.label) for c in _template_simple_components(db, template_id)]
    fields.append(ExcelImportTargetField(name=META_STATUS_FIELD, label="Estado (opcional)"))
    fields.append(ExcelImportTargetField(name=META_CREATED_AT_FIELD, label="Fecha histórica (opcional)"))
    return fields


def _record_aliases(db: Session, template_id: str) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for component in _template_simple_components(db, template_id):
        aliases[component.label.strip().lower()] = component.name
        aliases[component.name.strip().lower()] = component.name
    aliases["estado"] = META_STATUS_FIELD
    aliases["status"] = META_STATUS_FIELD
    aliases["fecha"] = META_CREATED_AT_FIELD
    aliases["fecha historica"] = META_CREATED_AT_FIELD
    aliases["fecha histórica"] = META_CREATED_AT_FIELD
    return aliases


def _to_read(db: Session, row: ExcelImportJob) -> ExcelImportJobRead:
    target_fields = _record_target_fields(db, row.template_id) if row.entity_type == "records" and row.template_id else None
    return ExcelImportJobRead(
        id=row.id,
        project_id=row.project_id,
        entity_type=row.entity_type,
        template_id=row.template_id,
        source_filename=row.source_filename,
        status=row.status,
        column_mapping=json.loads(row.column_mapping_json) if row.column_mapping_json else None,
        preview=json.loads(row.preview_json) if row.preview_json else None,
        target_fields=target_fields,
        total_rows=row.total_rows,
        imported_rows=row.imported_rows,
        failed_rows=row.failed_rows,
        error_report=json.loads(row.error_report_json) if row.error_report_json else None,
        created_at=row.created_at,
        completed_at=row.completed_at,
    )


class ExcelImportService:
    def upload_and_preview(self, db: Session, project_id: str, entity_type: str, filename: str, content: bytes, user_id: str, template_id: str | None = None) -> ExcelImportJobRead:
        if entity_type not in {*ENTITY_ALIASES.keys(), "records"}:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="Tipo de entidad no soportado para carga Excel")
        if entity_type == "records":
            if not template_id:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="template_id es obligatorio para cargar registros historicos")
            template = db.query(BuilderTemplate).filter(BuilderTemplate.id == template_id, BuilderTemplate.project_id == project_id).first()
            if template is None:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="La plantilla indicada no existe en este proyecto")
        else:
            template_id = None
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            first_row = next(rows_iter, [])
            headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=f"No fue posible leer el archivo Excel: {exc}") from exc

        data_rows: list[dict[str, object]] = []
        for values in rows_iter:
            if values is None or all(value is None for value in values):
                continue
            data_rows.append({headers[i]: values[i] for i in range(len(headers)) if i < len(values)})

        auto_mapping = self._auto_detect_mapping(db, headers, entity_type, template_id)

        row = ExcelImportJob(
            project_id=project_id,
            entity_type=entity_type,
            template_id=template_id,
            source_filename=filename,
            status="uploaded",
            column_mapping_json=json.dumps(auto_mapping),
            preview_json=json.dumps({"headers": headers, "sample_rows": data_rows[:PREVIEW_ROW_LIMIT]}),
            rows_json=json.dumps(data_rows),
            total_rows=len(data_rows),
            created_by=user_id,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return _to_read(db, row)

    def confirm_mapping(self, db: Session, job_id: str, column_mapping: dict[str, str]) -> ExcelImportJobRead:
        row = self._get_or_404(db, job_id)
        if row.status not in {"uploaded", "mapped"}:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="El lote ya fue procesado")
        row.column_mapping_json = json.dumps(column_mapping)
        row.status = "mapped"
        db.commit()
        db.refresh(row)
        return _to_read(db, row)

    def approve_and_import(self, db: Session, job_id: str, approved_by: str) -> ExcelImportJobRead:
        row = self._get_or_404(db, job_id)
        if row.status != "mapped":
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="El lote debe estar mapeado antes de aprobarse")

        mapping = json.loads(row.column_mapping_json) if row.column_mapping_json else {}
        data_rows = json.loads(row.rows_json) if row.rows_json else []
        required = REQUIRED_FIELDS.get(row.entity_type, set())
        if row.entity_type == "participants":
            target_fields = PARTICIPANT_TARGET_FIELDS
        elif row.entity_type == "assignments":
            target_fields = ASSIGNMENT_TARGET_FIELDS
        elif row.entity_type == "records":
            target_fields = {field.name for field in _record_target_fields(db, row.template_id)}
        else:
            target_fields = USER_TARGET_FIELDS

        errors: list[dict[str, object]] = []
        imported = 0
        for index, data_row in enumerate(data_rows):
            mapped_raw = {target_field: data_row.get(source_header) for source_header, target_field in mapping.items() if target_field in target_fields}
            mapped = {key: (str(value).strip() if value is not None else None) for key, value in mapped_raw.items()}
            missing = [field for field in required if not mapped.get(field)]
            if missing:
                errors.append({"row": index + 2, "error": f"Campos obligatorios faltantes: {', '.join(missing)}"})
                continue
            try:
                if row.entity_type == "participants":
                    participant_service.create_participant(db, ParticipantCreate(project_id=row.project_id, **mapped))
                elif row.entity_type == "assignments":
                    self._import_assignment_row(db, row.project_id, mapped)
                elif row.entity_type == "records":
                    self._import_record_row(db, row.project_id, row.template_id, mapped, approved_by)
                else:
                    identity_service.create_user(db, UserCreate(**mapped))
                imported += 1
            except HTTPException as exc:
                db.rollback()
                errors.append({"row": index + 2, "error": str(exc.detail)})
            except Exception as exc:
                db.rollback()
                errors.append({"row": index + 2, "error": str(exc)})

        row.status = "completed"
        row.imported_rows = imported
        row.failed_rows = len(errors)
        row.error_report_json = json.dumps(errors) if errors else None
        row.approved_by = approved_by
        row.approved_at = utc_now()
        row.completed_at = utc_now()
        db.commit()
        db.refresh(row)
        return _to_read(db, row)

    def _import_record_row(self, db: Session, project_id: str, template_id: str, mapped: dict[str, str | None], user_id: str) -> None:
        """Crea un RuntimeRecord historico reutilizando runtime_record_service.save_record.

        Reutilizar save_record (en vez de insertar RuntimeRecord/RuntimeRecordValue
        a mano) es lo que le da a esta carga masiva "el mismo peaje de validacion"
        que una captura real desde Runtime: enlace automatico de participante por
        DOCUMENT_ID (docs/98), deteccion de posible duplicado (docs/77), snapshot
        del flujo de aprobacion activo y consecutivos SERIAL_NUMBER si aplican.
        Ver docs/104.
        """
        components = {component.name: component for component in _template_simple_components(db, template_id)}

        status_value = (mapped.get(META_STATUS_FIELD) or "submitted").strip().lower()
        if status_value not in RECORD_VALID_STATUSES:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=f"Estado '{status_value}' no valido, use uno de: {', '.join(sorted(RECORD_VALID_STATUSES))}")

        created_at_override: datetime | None = None
        raw_created_at = mapped.get(META_CREATED_AT_FIELD)
        if raw_created_at:
            created_at_override = self._parse_historical_date(raw_created_at)

        values: list[RuntimeValueCreate] = []
        for field_name, raw_value in mapped.items():
            if field_name in {META_STATUS_FIELD, META_CREATED_AT_FIELD}:
                continue
            component = components.get(field_name)
            if component is None or not raw_value:
                continue
            values.append(RuntimeValueCreate(
                component_id=component.id,
                field_name=field_name,
                field_value_json=self._coerce_field_value(component.component_type, raw_value),
            ))

        if not values:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="La fila no tiene ningun valor de campo mapeado")

        try:
            saved = runtime_record_service.save_record(db, RuntimeRecordCreate(
                project_id=project_id,
                template_id=template_id,
                status=status_value,
                values=values,
            ), user_id)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(exc)) from exc

        if created_at_override is not None:
            record = db.query(RuntimeRecord).filter(RuntimeRecord.id == saved.id).first()
            record.created_at = created_at_override
            db.flush()

    def _coerce_field_value(self, component_type: str, raw_value: str) -> str:
        if component_type in RECORD_NUMERIC_FIELD_TYPES:
            try:
                number = float(raw_value.replace(",", "."))
            except ValueError as exc:
                raise ValueError(f"El valor '{raw_value}' no es un numero valido para un campo {component_type}") from exc
            if number.is_integer():
                number = int(number)
            return json.dumps(number)
        if component_type == "BOOLEAN":
            normalized = raw_value.strip().lower()
            if normalized in RECORD_BOOLEAN_TRUE_VALUES:
                return json.dumps(True)
            if normalized in RECORD_BOOLEAN_FALSE_VALUES:
                return json.dumps(False)
            raise ValueError(f"El valor '{raw_value}' no es un booleano reconocido (use si/no, true/false)")
        return json.dumps(raw_value)

    def _parse_historical_date(self, raw_value: str) -> datetime:
        candidates = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")
        for date_format in candidates:
            try:
                return datetime.strptime(raw_value.strip(), date_format)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(raw_value.strip())
        except ValueError as exc:
            raise ValueError(f"Fecha historica '{raw_value}' no reconocida (use AAAA-MM-DD)") from exc

    def _import_assignment_row(self, db: Session, project_id: str, mapped: dict[str, str | None]) -> None:
        """Asigna un usuario ya existente a `project_id` con el rol indicado.

        A diferencia de entity_type="users" (que crea usuarios nuevos), esto
        asume que el usuario y el rol ya existen -- carga masiva de
        asignaciones, no de identidades (ver docs/103). No hay deduplicacion:
        si la fila se importa dos veces, o el usuario ya tenia una asignacion
        activa en el proyecto, queda una fila UserProjectAssignment mas --
        mismo comportamiento que crear una asignacion individual hoy.
        """
        email = mapped["email"]
        user = identity_service.get_user_by_email(db, email)
        if user is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"No existe un usuario con el correo '{email}'")

        role_name = mapped["role_name"]
        role = db.query(Role).filter(func.lower(Role.name) == role_name.strip().lower()).first()
        if role is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"No existe un rol llamado '{role_name}'")

        assignment_service.create_assignment(db, AssignmentCreate(
            user_id=user.id,
            project_id=project_id,
            role_id=role.id,
            status=mapped.get("status") or "active",
        ))

    def get_job(self, db: Session, job_id: str) -> ExcelImportJobRead | None:
        row = db.query(ExcelImportJob).filter(ExcelImportJob.id == job_id).first()
        return _to_read(db, row) if row else None

    def list_jobs(self, db: Session, project_id: str) -> list[ExcelImportJobRead]:
        rows = db.query(ExcelImportJob).filter(ExcelImportJob.project_id == project_id).order_by(ExcelImportJob.created_at.desc()).all()
        return [_to_read(db, row) for row in rows]

    def _auto_detect_mapping(self, db: Session, headers: list[str], entity_type: str, template_id: str | None) -> dict[str, str]:
        aliases = _record_aliases(db, template_id) if entity_type == "records" else ENTITY_ALIASES[entity_type]
        mapping: dict[str, str] = {}
        for header in headers:
            target = aliases.get(header.strip().lower())
            if target:
                mapping[header] = target
        return mapping

    def _get_or_404(self, db: Session, job_id: str) -> ExcelImportJob:
        row = db.query(ExcelImportJob).filter(ExcelImportJob.id == job_id).first()
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote de carga Excel no encontrado")
        return row


excel_import_service = ExcelImportService()
