"""Detecta el formato de un archivo de formulario subido y lo despacha al
importador correspondiente, para que `/xlsform/import` acepte de forma
transparente XLSForm, o los formatos SurveyMonkey/LimeSurvey descritos en
`multi_format_import_service.py`.

Nota de honestidad (repetida aqui porque es la decision de diseno central
de este modulo): "surveymonkey"/"limesurvey" son los formatos definidos en
la plantilla de referencia del usuario, no necesariamente identicos a
cualquier exportacion real de esas plataformas -- ver docs/94.
"""

from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.schemas.xlsform import XlsformImportResult
from app.services import multi_format_import_service
from app.services.xlsform_import_service import xlsform_import_service

SURVEYMONKEY_HEADERS = {"identificador_pregunta", "texto_pregunta", "tipo_pregunta", "obligatorio"}
LIMESURVEY_HEADERS = {"questioncode", "questiontext", "questiontype", "isrequired"}


def _normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_") if value is not None else ""


def detect_format(content: bytes) -> str:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return "unknown"

    if "survey" in workbook.sheetnames:
        return "xlsform"

    first_sheet = workbook[workbook.sheetnames[0]]
    headers = {_normalize_header(cell) for cell in next(first_sheet.iter_rows(values_only=True), [])}
    if SURVEYMONKEY_HEADERS.issubset(headers):
        return "surveymonkey"
    if LIMESURVEY_HEADERS.issubset(headers):
        return "limesurvey"
    return "unknown"


def import_form(db: Session, project_id: str, filename: str, content: bytes, user_id: str | None, replace_template_id: str | None = None) -> XlsformImportResult:
    format_name = detect_format(content)
    if format_name == "xlsform":
        return xlsform_import_service.import_xlsform(db, project_id, filename, content, user_id, replace_template_id)
    if format_name == "surveymonkey":
        return multi_format_import_service.import_surveymonkey(db, project_id, filename, content, user_id, replace_template_id)
    if format_name == "limesurvey":
        return multi_format_import_service.import_limesurvey(db, project_id, filename, content, user_id, replace_template_id)
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
        detail=(
            "Formato de archivo no reconocido. Se acepta XLSForm (hoja 'survey' con columnas type/name), "
            "o el formato SurveyMonkey/LimeSurvey de la plantilla de referencia "
            "(columnas Identificador_Pregunta/Texto_Pregunta/Tipo_Pregunta/Obligatorio, "
            "o QuestionCode/QuestionText/QuestionType/IsRequired)."
        ),
    )
