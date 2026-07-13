"""Importadores para los formatos SurveyMonkey y LimeSurvey.

Estos dos formatos, a diferencia de XLSForm, no son un estandar publicado
por esas plataformas -- SurveyMonkey y LimeSurvey no ofrecen una
exportacion de *diseno* de encuesta en Excel con columnas fijas (solo
exportan *respuestas*, o en el caso de LimeSurvey, un formato propio
`.lss`/`.lsg` en XML). Las columnas que reconoce este importador
(`Identificador_Pregunta`/`Texto_Pregunta`/`Tipo_Pregunta`/... para
SurveyMonkey, `QuestionCode`/`QuestionText`/`QuestionType`/... para
LimeSurvey) son las definidas en la plantilla de referencia que aporto el
usuario (`plantilla_maestra_formularios_completa.xlsx`, hojas
`SurveyMonkey_Template` y `LiveSurvey_Template`) -- ver `docs/94` para el
detalle de este alcance y su limite conocido.

A diferencia de XLSForm (dos hojas, `survey`+`choices`, con listas de
opciones separadas), estos dos formatos van en una sola hoja con las
opciones embebidas en la misma fila como texto separado por comas o por
`|`. Por eso no comparten el parser de `xlsform_import_service.py` mas
alla de la creacion de componentes de Builder (`form_import_common`).
"""

import re
import unicodedata
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.schemas.builder_layout import BuilderPageCreate, BuilderSectionCreate
from app.schemas.xlsform import XlsformImportResult
from app.services.builder_layout_service import builder_layout_service
from app.services.form_import_common import create_field_component, prepare_target_template

TRUE_VALUES = {"yes", "sí", "si", "true", "1"}

# Tipo internos que NO deben interpretar su columna de "opciones" como una
# lista real de opciones -- en la plantilla de referencia esa columna trae
# texto descriptivo de la escala (p.ej. "Escala 1 a 5", "Min: 0, Max: 100"),
# no opciones literales.
SCALE_DESCRIPTOR_TYPES = {"NPS", "RATING"}
RANGE_PARAMS_RE = re.compile(r"min[:\s]*(-?\d+(?:\.\d+)?).*max[:\s]*(-?\d+(?:\.\d+)?)", re.IGNORECASE)

# label en minuscula (tal como viene en la columna Tipo_Pregunta) -> (tipo interno, appearance)
SURVEYMONKEY_TYPE_MAP: dict[str, tuple[str, str | None]] = {
    "opción múltiple (selección única)": ("SELECT", None),
    "casillas de verificación (selección múltiple)": ("MULTISELECT", None),
    "cuadro de texto de líneas múltiples": ("TEXTAREA", None),
    "cuadro de texto de una sola línea": ("TEXT", None),
    "matriz / escala de calificación": ("MATRIX", None),
    "clasificación / ranking": ("RANKING", None),
    "fecha / hora": ("DATETIME", None),
    "net promoter® score (nps)": ("NPS", None),
    "deslizador / slider": ("RANGE", None),
    "carga de archivos": ("FILE", None),
    "formulario de información de contacto": ("TEXT", None),
}

LIMESURVEY_TYPE_MAP: dict[str, tuple[str, str | None]] = {
    "short text (texto corto)": ("TEXT", None),
    "long text (texto largo/memo)": ("TEXTAREA", None),
    "radio button (selección única horizontal)": ("SELECT", "horizontal"),
    "dropdown list (menú desplegable)": ("DROPDOWN", None),
    "checkboxes (selección múltiple)": ("MULTISELECT", None),
    "number (numérico estricto)": ("INTEGER", None),
    "date picker (selector de fecha)": ("DATE", None),
    "file upload (carga de archivos)": ("FILE", None),
    "yes/no toggle (botón de alternancia)": ("BOOLEAN", None),
    "star rating (calificación con estrellas)": ("RATING", None),
    "consent checkbox (aceptación de términos)": ("BOOLEAN", "consent"),
}

SELECT_LIKE_TYPES = {"SELECT", "MULTISELECT", "DROPDOWN", "RANKING"}


def _normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_") if value is not None else ""


def _find_column(headers: list[str], name: str) -> int | None:
    return headers.index(name) if name in headers else None


def _cell(row: list[object], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _slugify(value: str, used: set[str]) -> str:
    normalized = unicodedata.normalize("NFD", value.strip().lower())
    ascii_only = "".join(character for character in normalized if not unicodedata.combining(character))
    base = re.sub(r"[^a-z0-9]+", "_", ascii_only).strip("_") or "campo"
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}_{index}"
        index += 1
    used.add(candidate)
    return candidate


def _read_first_sheet(content: bytes) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(cell) for cell in next(rows_iter, [])]
    rows = [list(row) for row in rows_iter if row and any(cell is not None for cell in row)]
    return headers, rows


def _build_template(db: Session, project_id: str, filename: str, source_label: str, parsed_fields: list[dict], replace_template_id: str | None) -> XlsformImportResult:
    template_id, is_replace = prepare_target_template(db, project_id, filename, replace_template_id)
    page = builder_layout_service.create_page(db, BuilderPageCreate(template_id=template_id, title=f"Importado de {source_label}", sort_order=0))
    section = builder_layout_service.create_section(db, BuilderSectionCreate(page_id=page.id, title="Preguntas", sort_order=0))

    imported_fields = 0
    warnings: list[str] = []
    used_names: set[str] = set()

    for index, item in enumerate(parsed_fields):
        name = _slugify(item["label"], used_names)
        config: dict = {}
        if item.get("required"):
            config["required"] = True
        if item.get("appearance"):
            config["appearance"] = item["appearance"]

        comp_type = item["type"]
        raw_options_text = item.get("raw_options_text") or ""
        if comp_type in SELECT_LIKE_TYPES and raw_options_text:
            separator = item["options_separator"]
            options = [chunk.strip() for chunk in raw_options_text.split(separator) if chunk.strip()]
            if options:
                config["options"] = [{"value": _slugify(option, set()), "label": option} for option in options]
        elif comp_type == "RANGE" and raw_options_text:
            match = RANGE_PARAMS_RE.search(raw_options_text)
            if match:
                config["min"] = float(match.group(1))
                config["max"] = float(match.group(2))
            else:
                warnings.append(f"Campo '{item['label']}': no se pudo interpretar el rango '{raw_options_text}'; se uso 0-100 por defecto")
        elif comp_type not in SCALE_DESCRIPTOR_TYPES and raw_options_text:
            # Tipo sin manejo especial de opciones pero con texto en esa
            # columna: se preserva como advertencia en vez de descartarlo en silencio.
            warnings.append(f"Campo '{item['label']}': el texto '{raw_options_text}' no se interpreto como opciones (tipo {comp_type})")

        if item.get("warning"):
            warnings.append(f"Campo '{item['label']}': {item['warning']}")

        create_field_component(db, template_id, section.id, index, component_type=comp_type, name=name, label=item["label"], config=config or None)
        imported_fields += 1

    return XlsformImportResult(template_id=template_id, imported_fields=imported_fields, warnings=warnings, replaced=is_replace)


def import_surveymonkey(db: Session, project_id: str, filename: str, content: bytes, user_id: str | None, replace_template_id: str | None = None) -> XlsformImportResult:
    headers, rows = _read_first_sheet(content)
    text_col = _find_column(headers, "texto_pregunta")
    type_col = _find_column(headers, "tipo_pregunta")
    options_col = _find_column(headers, "opciones_respuesta_separadas_por_comas")
    required_col = _find_column(headers, "obligatorio")
    if text_col is None or type_col is None:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="El archivo SurveyMonkey debe tener columnas 'Texto_Pregunta' y 'Tipo_Pregunta'")

    parsed = []
    for row in rows:
        label = _cell(row, text_col)
        if not label:
            continue
        raw_type = _cell(row, type_col).strip().lower()
        mapped = SURVEYMONKEY_TYPE_MAP.get(raw_type)
        warning = None
        if mapped is None:
            comp_type, appearance = "HIDDEN", None
            warning = f"tipo '{raw_type}' no reconocido en el formato SurveyMonkey; se importo como campo oculto"
        else:
            comp_type, appearance = mapped
            if raw_type == "formulario de información de contacto":
                warning = "campo de contacto compuesto (nombre/direccion/telefono) simplificado a un unico texto libre"
        parsed.append({
            "label": label, "type": comp_type, "appearance": appearance,
            "raw_options_text": _cell(row, options_col), "options_separator": ",",
            "required": _cell(row, required_col).strip().lower() in TRUE_VALUES,
            "warning": warning,
        })

    return _build_template(db, project_id, filename, "SurveyMonkey", parsed, replace_template_id)


def import_limesurvey(db: Session, project_id: str, filename: str, content: bytes, user_id: str | None, replace_template_id: str | None = None) -> XlsformImportResult:
    headers, rows = _read_first_sheet(content)
    text_col = _find_column(headers, "questiontext")
    type_col = _find_column(headers, "questiontype")
    options_col = _find_column(headers, "answerchoices_pipeseparated")
    required_col = _find_column(headers, "isrequired")
    if text_col is None or type_col is None:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="El archivo LimeSurvey debe tener columnas 'QuestionText' y 'QuestionType'")

    parsed = []
    for row in rows:
        label = _cell(row, text_col)
        if not label:
            continue
        raw_type = _cell(row, type_col).strip().lower()
        mapped = LIMESURVEY_TYPE_MAP.get(raw_type)
        warning = None
        if mapped is None:
            comp_type, appearance = "HIDDEN", None
            warning = f"tipo '{raw_type}' no reconocido en el formato LimeSurvey; se importo como campo oculto"
        else:
            comp_type, appearance = mapped
        parsed.append({
            "label": label, "type": comp_type, "appearance": appearance,
            "raw_options_text": _cell(row, options_col), "options_separator": "|",
            "required": _cell(row, required_col).strip().lower() in TRUE_VALUES,
            "warning": warning,
        })

    return _build_template(db, project_id, filename, "LimeSurvey", parsed, replace_template_id)
