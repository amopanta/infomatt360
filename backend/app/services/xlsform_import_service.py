"""Importador de plantillas XLSForm (estandar ODK/KoboToolbox).

Lee las hojas `survey` y `choices` de un archivo .xlsx y crea una plantilla
Builder equivalente, reutilizando el catalogo de tipos ya existente
(`app.core.field_types`), que ya conoce los alias de ODK (`select_one`,
`select_multiple`, `geopoint`, `begin_repeat`, etc.).

No requiere generar codigo nuevo de mapeo de tipos para la mayoria de casos:
`normalize_field_type()` ya acepta nombres XLSForm en minuscula porque los
convierte a mayuscula antes de comparar contra el catalogo. Solo se resuelven
aqui las particularidades propias del formato de archivo: `select_one <lista>`
con el nombre de lista embebido, grupos (`begin_group`/`end_group`, que se
aplanan) y repeticiones (`begin_repeat`/`end_repeat`, que se agrupan en un
solo componente REPEAT con los campos anidados en `config_json`).

Ademas de `type`/`name`/`label`, se leen las columnas reales del estandar
XLSForm que ya tienen un equivalente directo en la configuracion que produce
el constructor visual (`hint`, `required`, `relevant`, `constraint`,
`constraint_message`, `parameters`) -- ver `_apply_common_columns` y
`docs/93_EXPORTADOR_XLSFORM.md` para el detalle de que tan fiel es cada
traduccion.
"""

import json
import re
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.field_types import normalize_field_type
from app.schemas.builder import BuilderComponentCreate, BuilderTemplateCreate
from app.schemas.builder_layout import BuilderColumnCreate, BuilderPageCreate, BuilderRowCreate, BuilderSectionCreate
from app.schemas.xlsform import XlsformImportResult
from app.services.builder_layout_service import builder_layout_service
from app.services.builder_service import builder_service

# Filas de metadatos propios de ODK/KoboToolbox sin equivalente de campo
# visible: se omiten en vez de forzar un mapeo artificial.
SKIP_TYPES = {
    "note", "start", "end", "today", "deviceid", "subscriberid", "simserial",
    "username", "audit", "text-audit", "calculate_here", "background-audio",
}
GROUP_OPEN_TYPES = {"begin_group", "begin group", "begin_repeat", "begin repeat"}
GROUP_CLOSE_TYPES = {"end_group", "end group", "end_repeat", "end repeat"}
LIST_TYPES_WITH_EMBEDDED_LIST = {"select_one", "select_multiple", "select_one_from_file", "select_multiple_from_file", "rank"}

TRUE_VALUES = {"yes", "true", "1", "true()"}

_NOT_EMPTY_RE = re.compile(r"^\$\{(\w+)\}\s*!=\s*(''|\"\")$")
_EMPTY_RE = re.compile(r"^\$\{(\w+)\}\s*=\s*(''|\"\")$")
_STRING_LENGTH_NOT_EMPTY_RE = re.compile(r"^string-length\(\$\{(\w+)\}\)\s*>\s*0$")
_EQUALS_RE = re.compile(r"^\$\{(\w+)\}\s*=\s*'([^']*)'$|^\$\{(\w+)\}\s*=\s*\"([^\"]*)\"$")
_NOT_EQUALS_RE = re.compile(r"^\$\{(\w+)\}\s*!=\s*'([^']*)'$|^\$\{(\w+)\}\s*!=\s*\"([^\"]*)\"$")

_REGEX_CONSTRAINT_RE = re.compile(r"regex\(\.\s*,\s*'([^']*)'\)")
_STRING_LENGTH_MIN_RE = re.compile(r"string-length\(\.\)\s*>=\s*(\d+)")
_STRING_LENGTH_MAX_RE = re.compile(r"string-length\(\.\)\s*<=\s*(\d+)")
_NUMERIC_MIN_RE = re.compile(r"\.\s*>=\s*(-?\d+(?:\.\d+)?)")
_NUMERIC_MAX_RE = re.compile(r"\.\s*<=\s*(-?\d+(?:\.\d+)?)")
_NUMERIC_MIN_STRICT_RE = re.compile(r"\.\s*>\s*(-?\d+(?:\.\d+)?)")
_NUMERIC_MAX_STRICT_RE = re.compile(r"\.\s*<\s*(-?\d+(?:\.\d+)?)")


def _normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _find_column(headers: list[str], *candidates: str) -> int | None:
    normalized_candidates = [candidate.replace(" ", "_") for candidate in candidates]
    for index, header in enumerate(headers):
        if header.replace(" ", "_") in normalized_candidates:
            return index
    for index, header in enumerate(headers):
        if any(header.startswith(candidate) for candidate in normalized_candidates):
            return index
    return None


def _read_sheet_rows(workbook, sheet_name: str) -> tuple[list[str], list[list[object]]]:
    if sheet_name not in workbook.sheetnames:
        return [], []
    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(cell) for cell in next(rows_iter, [])]
    rows = [list(row) for row in rows_iter if row and any(cell is not None for cell in row)]
    return headers, rows


def _cell(row: list[object], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _parse_required(value: str) -> bool:
    return value.strip().lower() in TRUE_VALUES


def _parse_relevant_expression(expression: str) -> tuple[dict | None, str | None]:
    expression = expression.strip()
    if not expression:
        return None, None

    match = _NOT_EMPTY_RE.match(expression) or _STRING_LENGTH_NOT_EMPTY_RE.match(expression)
    if match:
        return {"field": match.group(1), "operator": "not_empty", "value": ""}, None
    match = _EMPTY_RE.match(expression)
    if match:
        return {"field": match.group(1), "operator": "empty", "value": ""}, None
    match = _EQUALS_RE.match(expression)
    if match:
        field, value = (match.group(1), match.group(2)) if match.group(1) else (match.group(3), match.group(4))
        return {"field": field, "operator": "equals", "value": value}, None
    match = _NOT_EQUALS_RE.match(expression)
    if match:
        field, value = (match.group(1), match.group(2)) if match.group(1) else (match.group(3), match.group(4))
        return {"field": field, "operator": "not_equals", "value": value}, None

    return None, f"expresion 'relevant' no reconocida, se preservo el texto pero no se activo como condicion: {expression!r}"


def _parse_constraint_expression(expression: str) -> tuple[dict, str | None]:
    expression = expression.strip()
    if not expression:
        return {}, None

    fields: dict[str, object] = {}
    remaining = expression
    for pattern, key, caster in (
        (_REGEX_CONSTRAINT_RE, "pattern", str),
        (_STRING_LENGTH_MIN_RE, "min_length", int),
        (_STRING_LENGTH_MAX_RE, "max_length", int),
        (_NUMERIC_MIN_RE, "min", float),
        (_NUMERIC_MAX_RE, "max", float),
        (_NUMERIC_MIN_STRICT_RE, "min", float),
        (_NUMERIC_MAX_STRICT_RE, "max", float),
    ):
        match = pattern.search(remaining)
        if match and key not in fields:
            fields[key] = caster(match.group(1))
            remaining = remaining[: match.start()] + remaining[match.end():]

    leftover = re.sub(r"\s+and\s+|\s+", " ", remaining).strip()
    warning = None
    if not fields or leftover:
        warning = f"expresion 'constraint' no se tradujo por completo a validaciones nativas, se preservo el texto original: {expression!r}"
    return fields, warning


def _parse_parameters(value: str) -> dict[str, str]:
    parts = re.split(r"[;,\s]+", value.strip())
    parsed: dict[str, str] = {}
    for part in parts:
        if "=" not in part:
            continue
        key, _, raw_value = part.partition("=")
        parsed[key.strip().lower()] = raw_value.strip()
    return parsed


class XlsformImportService:
    def import_xlsform(self, db: Session, project_id: str, filename: str, content: bytes, user_id: str | None) -> XlsformImportResult:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=f"No fue posible leer el archivo XLSForm: {exc}") from exc

        survey_headers, survey_rows = _read_sheet_rows(workbook, "survey")
        if not survey_rows:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="El archivo no tiene una hoja 'survey' con filas de preguntas")
        choices_headers, choices_rows = _read_sheet_rows(workbook, "choices")

        type_col = _find_column(survey_headers, "type")
        name_col = _find_column(survey_headers, "name")
        label_col = _find_column(survey_headers, "label")
        if type_col is None or name_col is None:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="La hoja 'survey' debe tener columnas 'type' y 'name'")

        hint_col = _find_column(survey_headers, "hint")
        required_col = _find_column(survey_headers, "required")
        relevant_col = _find_column(survey_headers, "relevant")
        constraint_col = _find_column(survey_headers, "constraint")
        constraint_message_col = _find_column(survey_headers, "constraint_message")
        appearance_col = _find_column(survey_headers, "appearance")
        parameters_col = _find_column(survey_headers, "parameters")

        list_col = _find_column(choices_headers, "list_name")
        choice_name_col = _find_column(choices_headers, "name")
        choice_label_col = _find_column(choices_headers, "label")
        choices_by_list: dict[str, list[dict[str, str]]] = {}
        for row in choices_rows:
            list_name = _cell(row, list_col)
            if not list_name:
                continue
            choices_by_list.setdefault(list_name, []).append({
                "value": _cell(row, choice_name_col),
                "label": _cell(row, choice_label_col) or _cell(row, choice_name_col),
            })

        template = builder_service.create_template(db, BuilderTemplateCreate(project_id=project_id, name=filename.rsplit(".", 1)[0], status="draft"))
        page = builder_layout_service.create_page(db, BuilderPageCreate(template_id=template.id, title="Importado de XLSForm", sort_order=0))
        section = builder_layout_service.create_section(db, BuilderSectionCreate(page_id=page.id, title="Preguntas", sort_order=0))

        warnings: list[str] = []
        imported_fields = 0
        sort_order = 0
        repeat_stack: list[dict[str, object]] = []

        for row in survey_rows:
            raw_type = _cell(row, type_col)
            if not raw_type:
                continue
            parts = raw_type.split()
            base_type = parts[0].lower()
            field_name = _cell(row, name_col)
            field_label = _cell(row, label_col) or field_name

            if base_type in ("begin_group", "begin group"):
                continue
            if base_type in ("end_group", "end group"):
                continue
            if base_type in ("begin_repeat", "begin repeat"):
                repeat_stack.append({"name": field_name, "label": field_label, "fields": []})
                continue
            if base_type in ("end_repeat", "end repeat"):
                if not repeat_stack:
                    warnings.append(f"'end_repeat' sin 'begin_repeat' correspondiente (fila con name={field_name!r})")
                    continue
                repeat = repeat_stack.pop()
                config = {"fields": repeat["fields"]}
                self._create_field_component(
                    db, template.id, section.id, sort_order,
                    component_type="REPEAT", name=str(repeat["name"]) or f"repeat_{sort_order}",
                    label=str(repeat["label"]) or "Repetible", config=config,
                )
                sort_order += 1
                imported_fields += 1
                continue
            if base_type in SKIP_TYPES:
                continue
            if not field_name:
                warnings.append(f"Fila con tipo '{raw_type}' sin columna 'name'; se omitio")
                continue

            mapped_type, config, warning = self._resolve_type(base_type, parts, choices_by_list)
            if warning:
                warnings.append(f"Campo '{field_name}': {warning}")

            config, common_warnings = self._apply_common_columns(
                config, mapped_type,
                hint=_cell(row, hint_col),
                required=_cell(row, required_col),
                relevant=_cell(row, relevant_col),
                constraint=_cell(row, constraint_col),
                constraint_message=_cell(row, constraint_message_col),
                appearance=_cell(row, appearance_col),
                parameters=_cell(row, parameters_col),
            )
            warnings.extend(f"Campo '{field_name}': {message}" for message in common_warnings)

            if repeat_stack:
                repeat_stack[-1]["fields"].append({"name": field_name, "label": field_label, "component_type": mapped_type, "config": config})
                continue

            self._create_field_component(db, template.id, section.id, sort_order, component_type=mapped_type, name=field_name, label=field_label, config=config)
            sort_order += 1
            imported_fields += 1

        if repeat_stack:
            warnings.append("El archivo termino con un 'begin_repeat' sin cerrar; los campos restantes se descartaron")

        return XlsformImportResult(template_id=template.id, imported_fields=imported_fields, warnings=warnings)

    def _resolve_type(self, base_type: str, parts: list[str], choices_by_list: dict[str, list[dict[str, str]]]) -> tuple[str, dict | None, str | None]:
        if base_type in LIST_TYPES_WITH_EMBEDDED_LIST:
            list_name = parts[1] if len(parts) > 1 else ""
            options = choices_by_list.get(list_name, [])
            config = {"options": options}
            if base_type in ("select_multiple", "select_multiple_from_file"):
                mapped = "MULTISELECT"
            elif base_type in ("select_one", "select_one_from_file"):
                mapped = "SELECT"
            else:
                mapped = "RANKING"
            warning = None if options or not list_name else f"lista de opciones '{list_name}' no encontrada en la hoja choices"
            return mapped, config, warning
        try:
            return normalize_field_type(base_type), None, None
        except ValueError:
            return "HIDDEN", None, f"tipo '{base_type}' no tiene equivalente directo; se importo como campo oculto"

    def _apply_common_columns(
        self, config: dict | None, mapped_type: str, *,
        hint: str, required: str, relevant: str, constraint: str,
        constraint_message: str, appearance: str, parameters: str,
    ) -> tuple[dict | None, list[str]]:
        """Traduce las columnas comunes de XLSForm a las claves de config_json
        que ya usa/entiende el constructor visual (placeholder, required,
        relevant, pattern/min/max/min_length/max_length). No asume que
        `config` ya sea un dict: la mayoria de tipos simples (TEXT, INTEGER,
        DATE...) llegan aqui con `config=None` desde `_resolve_type`, y estas
        columnas son justamente las que mas les aplican. Devuelve el config
        actualizado (o el original si no habia nada que agregar) y
        advertencias legibles."""
        warnings: list[str] = []
        updates: dict[str, object] = {}

        if hint:
            updates["placeholder"] = hint
        if required:
            updates["required"] = _parse_required(required)
        if appearance:
            updates["appearance"] = appearance

        if relevant:
            updates["relevant_expression"] = relevant
            parsed_relevant, warning = _parse_relevant_expression(relevant)
            if parsed_relevant:
                updates["relevant"] = parsed_relevant
            if warning:
                warnings.append(warning)

        if constraint:
            updates["constraint_expression"] = constraint
            parsed_constraint, warning = _parse_constraint_expression(constraint)
            updates.update(parsed_constraint)
            if warning:
                warnings.append(warning)
        if constraint_message:
            updates["constraint_message"] = constraint_message

        if parameters and mapped_type == "RANGE":
            parsed_parameters = _parse_parameters(parameters)
            if "start" in parsed_parameters:
                updates["min"] = float(parsed_parameters["start"])
            if "end" in parsed_parameters:
                updates["max"] = float(parsed_parameters["end"])
            if "step" in parsed_parameters:
                updates["step"] = float(parsed_parameters["step"])
        elif parameters:
            updates["parameters"] = parameters

        if not updates:
            return config, warnings
        merged = dict(config or {})
        merged.update(updates)
        return merged, warnings

    def _create_field_component(self, db: Session, template_id: str, section_id: str, sort_order: int, *, component_type: str, name: str, label: str, config: dict | None) -> None:
        row = builder_layout_service.create_row(db, BuilderRowCreate(section_id=section_id, sort_order=sort_order))
        column = builder_layout_service.create_column(db, BuilderColumnCreate(row_id=row.id, desktop_width=12, tablet_width=12, mobile_width=12, sort_order=0))
        builder_service.add_component(db, BuilderComponentCreate(
            template_id=template_id,
            column_id=column.id,
            component_type=component_type,
            name=name,
            label=label,
            config_json=json.dumps(config) if config is not None else None,
            sort_order=sort_order,
        ))


xlsform_import_service = XlsformImportService()
