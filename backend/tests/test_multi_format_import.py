"""Pruebas del importador multi-formato (SurveyMonkey/LimeSurvey), usando
las filas EXACTAS de la plantilla de referencia del usuario
(`plantilla_maestra_formularios_completa.xlsx`, hojas `SurveyMonkey_Template`
y `LiveSurvey_Template`) para verificar fidelidad real, no datos inventados.
"""

import json
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.testclient import TestClient

from app.core.security import hash_password
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.assignment import UserProjectAssignment
from app.models.builder import BuilderComponent, BuilderTemplate
from app.models.identity import Project, Role, User
from app.services.form_import_router import detect_format


def _build_workbook(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preguntas"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


SURVEYMONKEY_HEADERS = ["Identificador_Pregunta", "Texto_Pregunta", "Tipo_Pregunta", "Opciones_Respuesta_Separadas_Por_Comas", "Obligatorio"]
SURVEYMONKEY_ROWS = [
    ["SM_01", "¿Cuál es su nivel de satisfacción?", "Opción Múltiple (Selección Única)", "Muy Satisfecho, Satisfecho, Insatisfecho", "Sí"],
    ["SM_02", "¿Qué dispositivos utiliza a diario?", "Casillas de Verificación (Selección Múltiple)", "Smartphone, Laptop, Tablet, Ninguno", "No"],
    ["SM_03", "Por favor, detalle su opinión aquí:", "Cuadro de Texto de Líneas Múltiples", None, "No"],
    ["SM_04", "Escriba su correo electrónico:", "Cuadro de Texto de una Sola Línea", None, "Sí"],
    ["SM_05", "Evalúe las siguientes características:", "Matriz / Escala de Calificación", "Filas: Diseño, Usabilidad, Soporte | Columnas: Malo, Regular, Bueno", "Sí"],
    ["SM_06", "Ordene de 1 a 3 sus prioridades:", "Clasificación / Ranking", "Opción A, Opción B, Opción C", "No"],
    ["SM_07", "Indique su fecha de ingreso:", "Fecha / Hora", None, "No"],
    ["SM_08", "¿Cuánto califica nuestro servicio del 1 al 10?", "Net Promoter® Score (NPS)", "Escala estándar 0-10", "Sí"],
    ["SM_09", "Ingrese un valor numérico exacto:", "Deslizador / Slider", "Min: 0, Max: 100", "No"],
    ["SM_10", "Suba el comprobante correspondiente:", "Carga de Archivos", "PDF, JPG, PNG", "No"],
    ["SM_11", "Información de contacto estructurada:", "Formulario de Información de Contacto", "Campos predefinidos: Nombre, Dirección, Teléfono", "Sí"],
]

LIMESURVEY_HEADERS = ["QuestionCode", "QuestionText", "QuestionType", "AnswerChoices_PipeSeparated", "IsRequired"]
LIMESURVEY_ROWS = [
    ["LS_01", "Ingrese sus comentarios generales:", "Short Text (Texto Corto)", None, "yes"],
    ["LS_02", "Describa detalladamente el incidente:", "Long Text (Texto Largo/Memo)", None, "no"],
    ["LS_03", "Seleccione su género:", "Radio Button (Selección Única Horizontal)", "Masculino|Femenino|Otro|Prefiero no decirlo", "yes"],
    ["LS_04", "Elija su país de residencia:", "Dropdown List (Menú Desplegable)", "Colombia|México|Perú|Argentina|Ecuador", "yes"],
    ["LS_05", "Marque sus pasatiempos favoritos:", "Checkboxes (Selección Múltiple)", "Lectura|Deportes|Cine|Música|Viajes", "no"],
    ["LS_06", "Ingrese únicamente valores enteros:", "Number (Numérico Estricto)", None, "no"],
    ["LS_07", "Seleccione el día del evento:", "Date Picker (Selector de Fecha)", None, "no"],
    ["LS_08", "Suba la imagen requerida:", "File Upload (Carga de Archivos)", "Límite 10MB", "no"],
    ["LS_09", "¿Recomendaría este servicio?", "Yes/No Toggle (Botón de Alternancia)", "Sí|No", "yes"],
    ["LS_10", "Califique la atención recibida:", "Star Rating (Calificación con Estrellas)", "Escala 1 a 5", "yes"],
    ["LS_11", "Bloque de confirmación legal:", "Consent Checkbox (Aceptación de Términos)", "Acepto los términos y condiciones de tratamiento de datos", "yes"],
]


def setup_client():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    sessions = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    with sessions() as db:
        project = Project(id="multiformat-project", name="Multi-formato")
        builder_role = Role(id="multiformat-builder-role", name="Builder", permissions="builder.write")
        builder = User(id="multiformat-builder", full_name="Builder", document_id="multiformat-builder-doc", email="multiformat-builder@example.com", password_hash=hash_password("Builder12345!"))
        db.add_all([
            project, builder_role, builder,
            UserProjectAssignment(user_id=builder.id, project_id=project.id, role_id=builder_role.id, status="active"),
        ])
        db.commit()

    def override_db():
        with sessions() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    return engine, sessions


def auth(client: TestClient, email: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_detect_format_recognizes_the_three_shapes():
    workbook = Workbook()
    workbook.active.title = "survey"
    workbook.active.append(["type", "name", "label"])
    buffer = BytesIO()
    workbook.save(buffer)
    assert detect_format(buffer.getvalue()) == "xlsform"
    assert detect_format(_build_workbook(SURVEYMONKEY_HEADERS, SURVEYMONKEY_ROWS)) == "surveymonkey"
    assert detect_format(_build_workbook(LIMESURVEY_HEADERS, LIMESURVEY_ROWS)) == "limesurvey"
    assert detect_format(_build_workbook(["a", "b", "c"], [[1, 2, 3]])) == "unknown"


def test_import_surveymonkey_reference_template_maps_every_row():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "multiformat-builder@example.com", "Builder12345!")
            content = _build_workbook(SURVEYMONKEY_HEADERS, SURVEYMONKEY_ROWS)

            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "multiformat-project"},
                files={"upload": ("encuesta_surveymonkey.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 200, response.text
            body = response.json()
            assert body["imported_fields"] == 11
            # SM_05 (matriz), SM_10 (tipos de archivo en la columna de opciones) y
            # SM_11 (contacto compuesto, doble advertencia) generan advertencias esperadas.
            assert len(body["warnings"]) == 4
            assert any("contacto compuesto" in warning for warning in body["warnings"])

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                by_label = {component.label: component for component in components}

                satisfaccion = by_label["¿Cuál es su nivel de satisfacción?"]
                assert satisfaccion.component_type == "SELECT"
                options = json.loads(satisfaccion.config_json)["options"]
                assert {option["label"] for option in options} == {"Muy Satisfecho", "Satisfecho", "Insatisfecho"}
                assert json.loads(satisfaccion.config_json)["required"] is True

                dispositivos = by_label["¿Qué dispositivos utiliza a diario?"]
                assert dispositivos.component_type == "MULTISELECT"

                correo = by_label["Escriba su correo electrónico:"]
                assert correo.component_type == "TEXT"

                deslizador = by_label["Ingrese un valor numérico exacto:"]
                assert deslizador.component_type == "RANGE"
                range_config = json.loads(deslizador.config_json)
                assert range_config["min"] == 0.0
                assert range_config["max"] == 100.0

                nps = by_label["¿Cuánto califica nuestro servicio del 1 al 10?"]
                assert nps.component_type == "NPS"
                # NPS es un descriptor de escala, no debe intentar parsear "options"
                assert "options" not in (json.loads(nps.config_json) if nps.config_json else {})

                contacto = by_label["Información de contacto estructurada:"]
                assert contacto.component_type == "TEXT"
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_import_limesurvey_reference_template_maps_every_row():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "multiformat-builder@example.com", "Builder12345!")
            content = _build_workbook(LIMESURVEY_HEADERS, LIMESURVEY_ROWS)

            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "multiformat-project"},
                files={"upload": ("encuesta_limesurvey.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 200, response.text
            body = response.json()
            assert body["imported_fields"] == 11

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                by_label = {component.label: component for component in components}

                genero = by_label["Seleccione su género:"]
                assert genero.component_type == "SELECT"
                genero_config = json.loads(genero.config_json)
                assert genero_config["appearance"] == "horizontal"
                assert {option["label"] for option in genero_config["options"]} == {"Masculino", "Femenino", "Otro", "Prefiero no decirlo"}

                pais = by_label["Elija su país de residencia:"]
                assert pais.component_type == "DROPDOWN"

                recomendaria = by_label["¿Recomendaría este servicio?"]
                assert recomendaria.component_type == "BOOLEAN"

                calificacion = by_label["Califique la atención recibida:"]
                assert calificacion.component_type == "RATING"

                consentimiento = by_label["Bloque de confirmación legal:"]
                assert consentimiento.component_type == "BOOLEAN"
                assert json.loads(consentimiento.config_json)["appearance"] == "consent"
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_unrecognized_format_returns_422_with_helpful_message():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "multiformat-builder@example.com", "Builder12345!")
            content = _build_workbook(["columna_rara_1", "columna_rara_2"], [["x", "y"]])

            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "multiformat-project"},
                files={"upload": ("desconocido.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 422
            assert "no reconocido" in response.json()["detail"]
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_combined_form_mixes_field_types_from_every_system_freely():
    """Demuestra que no hay silos por 'sistema': un mismo formulario puede
    combinar libremente un campo repetible (Kobo), una matriz (SurveyMonkey),
    un consentimiento (LimeSurvey) y un campo tipo referencia (ActivityInfo),
    porque todos son simplemente BuilderComponent en la misma plantilla."""
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "multiformat-builder@example.com", "Builder12345!")
            workbook = Workbook()
            survey = workbook.active
            survey.title = "survey"
            survey.append(["type", "name", "label", "hint", "required"])
            for row in [
                ["begin_repeat", "integrantes", "Integrantes del hogar", "", ""],
                ["text", "nombre_integrante", "Nombre", "", ""],
                ["end_repeat", "integrantes", "", "", ""],
                ["matrix", "evaluacion", "Evaluacion de servicio", "", ""],
                ["acknowledge", "consiento", "Acepto el tratamiento de datos", "", "yes"],
                ["reference", "programa_relacionado", "Programa relacionado", "", ""],
                ["range", "satisfaccion", "Satisfaccion", "", ""],
            ]:
                survey.append(row)
            buffer = BytesIO()
            workbook.save(buffer)
            content = buffer.getvalue()

            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "multiformat-project"},
                files={"upload": ("combinado.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 200, response.text
            body = response.json()
            assert body["imported_fields"] == 5  # repeat + matrix + acknowledge + reference + range

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                types_by_name = {component.name: component.component_type for component in components}
                assert types_by_name["integrantes"] == "REPEAT"
                assert types_by_name["evaluacion"] == "MATRIX"
                assert types_by_name["consiento"] == "BOOLEAN"  # acknowledge (LimeSurvey) es alias de BOOLEAN
                assert types_by_name["programa_relacionado"] == "REFERENCE"
                assert types_by_name["satisfaccion"] == "RANGE"
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)
