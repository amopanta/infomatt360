"""Pruebas del reemplazo de plantilla en el mismo lugar (mismo `template_id`,
como el "redeploy" de KoboToolbox), disponible para los 3 importadores
(XLSForm, SurveyMonkey, LimeSurvey) via `replace_template_id` en
`POST /xlsform/import`.
"""

import json
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.testclient import TestClient

from app.core.security import hash_password
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.assignment import UserProjectAssignment
from app.models.builder import BuilderComponent, BuilderTemplate, BuilderVersion
from app.models.identity import Project, Role, User
from app.models.runtime_record import RuntimeRecord


def _build_xlsform(rows: list[list[object]], choices: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(["type", "name", "label"])
    for row in rows:
        survey.append(row)
    choices_sheet = workbook.create_sheet("choices")
    choices_sheet.append(["list_name", "name", "label"])
    for row in choices or []:
        choices_sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def setup_client():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    sessions = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    with sessions() as db:
        project = Project(id="replace-project", name="Reemplazo de plantillas")
        other_project = Project(id="replace-other-project", name="Otro proyecto")
        builder_role = Role(id="replace-builder-role", name="Disenador", permissions="builder.write")
        builder = User(id="replace-builder", full_name="Disenador", document_id="replace-builder-doc", email="replace-builder@example.com", password_hash=hash_password("Builder12345!"))

        original_template = BuilderTemplate(id="replace-template", project_id=project.id, name="Censo original", status="published")

        db.add_all([
            project, other_project, builder_role, builder, original_template,
            UserProjectAssignment(user_id=builder.id, project_id=project.id, role_id=builder_role.id, status="active"),
            UserProjectAssignment(user_id=builder.id, project_id=other_project.id, role_id=builder_role.id, status="active"),
        ])
        db.commit()

    def override_db():
        with sessions() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    return engine, sessions


def auth(client: TestClient, email: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_import_without_replace_template_id_still_creates_a_new_template():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "replace-builder@example.com", "Builder12345!")
            content = _build_xlsform([["text", "nombre", "Nombre"]])
            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project"},
                files={"upload": ("nuevo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 200, response.text
            body = response.json()
            assert body["replaced"] is False
            assert body["template_id"] != "replace-template"
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_replace_keeps_same_template_id_and_swaps_fields():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "replace-builder@example.com", "Builder12345!")

            # Version original: nombre, edad.
            original = _build_xlsform([["text", "nombre", "Nombre"], ["integer", "edad", "Edad"]])
            created = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project"},
                files={"upload": ("censo_v1.xlsx", original, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert created.status_code == 200, created.text
            template_id = created.json()["template_id"]

            # Un registro capturado con la version original -- debe sobrevivir al reemplazo.
            with sessions() as db:
                db.add(RuntimeRecord(project_id="replace-project", template_id=template_id, status="submitted"))
                db.commit()
                record_id = db.query(RuntimeRecord).filter(RuntimeRecord.template_id == template_id).one().id

            # Version nueva: reemplaza nombre/edad por email/telefono, en el MISMO template_id.
            replacement = _build_xlsform([["text", "correo", "Correo"], ["text", "telefono", "Telefono"]])
            replaced = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project", "replace_template_id": template_id},
                files={"upload": ("censo_v2.xlsx", replacement, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert replaced.status_code == 200, replaced.text
            body = replaced.json()
            assert body["replaced"] is True
            assert body["template_id"] == template_id  # mismo lugar, no una plantilla nueva
            assert body["imported_fields"] == 2

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == template_id).all()
                names = {component.name for component in components}
                assert names == {"correo", "telefono"}  # nombre/edad ya no estan

                # El registro capturado antes del reemplazo sigue existiendo, intacto.
                assert db.query(RuntimeRecord).filter(RuntimeRecord.id == record_id).one() is not None

                # Se guardo un respaldo de la estructura anterior antes de sobrescribir.
                versions = db.query(BuilderVersion).filter(BuilderVersion.template_id == template_id).all()
                assert len(versions) == 1
                assert versions[0].status == "archived"
                snapshot = json.loads(versions[0].schema_json)
                snapshot_field_names = {
                    component["name"]
                    for page in snapshot["pages"]
                    for section in page["sections"]
                    for row in section["rows"]
                    for column in row["columns"]
                    for component in column["components"]
                }
                assert snapshot_field_names == {"nombre", "edad"}  # la version vieja quedo preservada
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_replace_rejects_template_from_another_project():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "replace-builder@example.com", "Builder12345!")
            content = _build_xlsform([["text", "nombre", "Nombre"]])
            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-other-project", "replace_template_id": "replace-template"},
                files={"upload": ("intento.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 404
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_replace_rejects_nonexistent_template():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "replace-builder@example.com", "Builder12345!")
            content = _build_xlsform([["text", "nombre", "Nombre"]])
            response = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project", "replace_template_id": "does-not-exist"},
                files={"upload": ("intento.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 404
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_replace_works_for_surveymonkey_format_too():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            headers = auth(client, "replace-builder@example.com", "Builder12345!")

            original = _build_xlsform([["text", "nombre", "Nombre"]])
            created = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project"},
                files={"upload": ("original.xlsx", original, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            template_id = created.json()["template_id"]

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "SurveyMonkey"
            sheet.append(["Identificador_Pregunta", "Texto_Pregunta", "Tipo_Pregunta", "Opciones_Respuesta_Separadas_Por_Comas", "Obligatorio"])
            sheet.append(["SM_01", "Nivel de satisfaccion", "Opción Múltiple (Selección Única)", "Bueno, Malo", "Sí"])
            buffer = BytesIO()
            workbook.save(buffer)

            replaced = client.post(
                "/api/v1/xlsform/import",
                headers=headers,
                data={"project_id": "replace-project", "replace_template_id": template_id},
                files={"upload": ("encuesta_sm.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert replaced.status_code == 200, replaced.text
            body = replaced.json()
            assert body["replaced"] is True
            assert body["template_id"] == template_id
            assert body["imported_fields"] == 1
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)
