import json
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.testclient import TestClient

from app.core.security import hash_password
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.assignment import UserProjectAssignment
from app.models.builder import BuilderComponent, BuilderTemplate
from app.models.builder_layout import BuilderColumn, BuilderPage, BuilderRow, BuilderSection
from app.models.identity import Project, Role, User


def setup_client():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    sessions = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    with sessions() as db:
        project = Project(id="xf-project", name="Exportacion XLSForm")
        builder_role = Role(id="xf-builder-role", name="Disenador", permissions="builder.write")
        outsider_role = Role(id="xf-outsider-role", name="Sin permiso", permissions="records.read")
        builder_user = User(id="xf-builder", full_name="Disenador", document_id="xf-builder-doc", email="xf-builder@example.com", password_hash=hash_password("Builder12345!"))
        outsider = User(id="xf-outsider", full_name="Sin permiso", document_id="xf-outsider-doc", email="xf-outsider@example.com", password_hash=hash_password("Outsider12345!"))

        template = BuilderTemplate(id="xf-template", project_id=project.id, name="Encuesta Feria", status="published")
        page = BuilderPage(id="xf-page", template_id=template.id, title="Pagina 1", sort_order=1)
        section = BuilderSection(id="xf-section", page_id=page.id, title="Datos", sort_order=1)

        rows_and_columns = []
        components = []
        for index, (comp_id, comp_type, name, label, config) in enumerate([
            ("xf-comp-text", "TEXT", "nombre_completo", "Nombre completo", None),
            ("xf-comp-select", "SELECT", "genero", "Genero", {"options": [{"value": "m", "label": "Masculino"}, {"value": "f", "label": "Femenino"}]}),
            ("xf-comp-bool", "BOOLEAN", "acepta_terminos", "Acepta terminos", None),
            ("xf-comp-repeat", "REPEAT", "integrantes_hogar", "Integrantes del hogar", {"fields": [{"name": "nombre_integrante", "label": "Nombre", "component_type": "TEXT", "config": {}}]}),
        ]):
            row = BuilderRow(id=f"xf-row-{index}", section_id=section.id, sort_order=index)
            column = BuilderColumn(id=f"xf-column-{index}", row_id=row.id, desktop_width=12, sort_order=1)
            rows_and_columns.extend([row, column])
            components.append(BuilderComponent(
                id=comp_id, template_id=template.id, column_id=column.id, component_type=comp_type,
                name=name, label=label, config_json=json.dumps(config) if config is not None else None, sort_order=index,
            ))

        rich_template = BuilderTemplate(id="xf-template-rich", project_id=project.id, name="Encuesta Rica", status="published")
        rich_page = BuilderPage(id="xf-rich-page", template_id=rich_template.id, title="Pagina 1", sort_order=1)
        rich_section = BuilderSection(id="xf-rich-section", page_id=rich_page.id, title="Datos", sort_order=1)
        rich_rows_and_columns = []
        rich_components = []
        for index, (comp_id, comp_type, name, label, config) in enumerate([
            ("xf-rich-comp-hint", "TEXT", "nombre_hint", "Nombre", {"placeholder": "Ejemplo: Juan Perez", "required": True}),
            ("xf-rich-comp-constraint", "INTEGER", "edad_validada", "Edad", {"min": 0, "max": 120}),
            ("xf-rich-comp-relevant", "TEXT", "comentario_condicional", "Comentario", {"relevant": {"field": "nombre_hint", "operator": "not_empty", "value": ""}}),
            ("xf-rich-comp-range", "RANGE", "satisfaccion", "Satisfaccion", {"min": 0, "max": 10, "step": 1}),
            ("xf-rich-comp-rank", "RANKING", "prioridades", "Prioridades", {"options": [{"value": "salud", "label": "Salud"}, {"value": "educacion", "label": "Educacion"}]}),
        ]):
            row = BuilderRow(id=f"xf-rich-row-{index}", section_id=rich_section.id, sort_order=index)
            column = BuilderColumn(id=f"xf-rich-column-{index}", row_id=row.id, desktop_width=12, sort_order=1)
            rich_rows_and_columns.extend([row, column])
            rich_components.append(BuilderComponent(
                id=comp_id, template_id=rich_template.id, column_id=column.id, component_type=comp_type,
                name=name, label=label, config_json=json.dumps(config) if config is not None else None, sort_order=index,
            ))

        db.add_all([
            project, builder_role, outsider_role, builder_user, outsider,
            template, page, section, *rows_and_columns, *components,
            rich_template, rich_page, rich_section, *rich_rows_and_columns, *rich_components,
            UserProjectAssignment(user_id=builder_user.id, project_id=project.id, role_id=builder_role.id, status="active"),
            UserProjectAssignment(user_id=outsider.id, project_id=project.id, role_id=outsider_role.id, status="active"),
        ])
        db.commit()

    def override_db():
        with sessions() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    return engine, sessions


def auth(client: TestClient, email: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_export_requires_builder_write_permission():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            outsider_headers = auth(client, "xf-outsider@example.com", "Outsider12345!")
            denied = client.get("/api/v1/xlsform/export/xf-template", headers=outsider_headers)
            assert denied.status_code == 403

            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            allowed = client.get("/api/v1/xlsform/export/xf-template", headers=builder_headers)
            assert allowed.status_code == 200
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_export_unknown_template_returns_404():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            response = client.get("/api/v1/xlsform/export/does-not-exist", headers=builder_headers)
            assert response.status_code == 404
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_export_produces_valid_xlsform_workbook_with_choices_and_repeat():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            response = client.get("/api/v1/xlsform/export/xf-template", headers=builder_headers)
            assert response.status_code == 200
            assert response.headers["content-disposition"].startswith('attachment; filename="Encuesta_Feria.xlsx"') or "Encuesta" in response.headers["content-disposition"]

            workbook = load_workbook(BytesIO(response.content))
            survey_rows = [tuple(row) for row in workbook["survey"].iter_rows(values_only=True)]
            choices_rows = [tuple(row) for row in workbook["choices"].iter_rows(values_only=True)]

            assert survey_rows[0] == ("type", "name", "label", "hint", "required", "relevant", "constraint", "constraint_message", "appearance", "parameters")
            survey_by_name = {row[1]: row for row in survey_rows[1:] if row[1]}
            assert survey_by_name["nombre_completo"][0] == "text"
            assert survey_by_name["genero"][0] == "select_one genero"
            assert survey_by_name["acepta_terminos"][0] == "select_one yes_no"

            # begin_repeat/end_repeat con el campo anidado en medio
            repeat_types = [row[0] for row in survey_rows[1:]]
            assert "begin_repeat" in repeat_types
            assert "end_repeat" in repeat_types
            assert any(row[1] == "nombre_integrante" for row in survey_rows[1:])

            choices_by_list: dict[str, list[tuple]] = {}
            for row in choices_rows[1:]:
                if not row[0]:
                    continue
                choices_by_list.setdefault(row[0], []).append(row)
            assert {row[1] for row in choices_by_list["genero"]} == {"m", "f"}
            assert {row[1] for row in choices_by_list["yes_no"]} == {"1", "0"}
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_exported_workbook_can_be_reimported_round_trip():
    """El archivo exportado debe ser un XLSForm valido: re-importarlo debe
    producir una plantilla nueva con los mismos campos basicos."""
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            exported = client.get("/api/v1/xlsform/export/xf-template", headers=builder_headers)
            assert exported.status_code == 200

            reimported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xf-project"},
                files={"upload": ("roundtrip.xlsx", exported.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert reimported.status_code == 200, reimported.text
            body = reimported.json()
            assert body["imported_fields"] == 4  # texto, select, boolean(select_one yes_no), repeat
            assert body["warnings"] == []

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                names = {component.name for component in components}
                assert "nombre_completo" in names
                assert "genero" in names
                assert "acepta_terminos" in names
                assert "integrantes_hogar" in names
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_export_writes_hint_required_relevant_constraint_range_and_rank_columns():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            response = client.get("/api/v1/xlsform/export/xf-template-rich", headers=builder_headers)
            assert response.status_code == 200

            workbook = load_workbook(BytesIO(response.content))
            survey_rows = [tuple(row) for row in workbook["survey"].iter_rows(values_only=True)]
            survey_by_name = {row[1]: row for row in survey_rows[1:] if row[1]}

            hint_row = survey_by_name["nombre_hint"]
            assert hint_row[3] == "Ejemplo: Juan Perez"  # hint
            assert hint_row[4] == "yes"  # required

            constraint_row = survey_by_name["edad_validada"]
            assert constraint_row[6] == ". >= 0 and . <= 120"  # constraint

            relevant_row = survey_by_name["comentario_condicional"]
            assert relevant_row[5] == "${nombre_hint} != ''"  # relevant

            range_row = survey_by_name["satisfaccion"]
            assert range_row[0] == "range"
            assert range_row[9] == "start=0 end=10 step=1"  # parameters

            rank_row = survey_by_name["prioridades"]
            assert rank_row[0] == "rank prioridades"

            choices_rows = [tuple(row) for row in workbook["choices"].iter_rows(values_only=True)]
            rank_choices = {row[1] for row in choices_rows[1:] if row[0] == "prioridades"}
            assert rank_choices == {"salud", "educacion"}
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_export_rich_workbook_round_trips_through_import():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            exported = client.get("/api/v1/xlsform/export/xf-template-rich", headers=builder_headers)
            assert exported.status_code == 200

            reimported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xf-project"},
                files={"upload": ("rich-roundtrip.xlsx", exported.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert reimported.status_code == 200, reimported.text
            body = reimported.json()
            assert body["imported_fields"] == 5
            assert body["warnings"] == []

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                by_name = {component.name: component for component in components}

                hint_config = json.loads(by_name["nombre_hint"].config_json)
                assert hint_config["placeholder"] == "Ejemplo: Juan Perez"
                assert hint_config["required"] is True

                constraint_config = json.loads(by_name["edad_validada"].config_json)
                assert constraint_config["min"] == 0.0
                assert constraint_config["max"] == 120.0

                relevant_config = json.loads(by_name["comentario_condicional"].config_json)
                assert relevant_config["relevant"] == {"field": "nombre_hint", "operator": "not_empty", "value": ""}

                assert by_name["satisfaccion"].component_type == "RANGE"
                range_config = json.loads(by_name["satisfaccion"].config_json)
                assert range_config["min"] == 0.0
                assert range_config["max"] == 10.0
                assert range_config["step"] == 1.0

                assert by_name["prioridades"].component_type == "RANKING"
                rank_config = json.loads(by_name["prioridades"].config_json)
                assert {option["value"] for option in rank_config["options"]} == {"salud", "educacion"}
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_master_template_requires_permission_and_covers_every_field_type():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            outsider_headers = auth(client, "xf-outsider@example.com", "Outsider12345!")
            denied = client.get("/api/v1/xlsform/master-template", headers=outsider_headers, params={"project_id": "xf-project"})
            assert denied.status_code == 403

            builder_headers = auth(client, "xf-builder@example.com", "Builder12345!")
            response = client.get("/api/v1/xlsform/master-template", headers=builder_headers, params={"project_id": "xf-project"})
            assert response.status_code == 200
            assert "plantilla_maestra_infomatt360.xlsx" in response.headers["content-disposition"]

            workbook = load_workbook(BytesIO(response.content))
            survey_rows = [tuple(row) for row in workbook["survey"].iter_rows(values_only=True)]
            assert survey_rows[0][:3] == ("type", "name", "label")
            assert len(survey_rows) > 40  # un ejemplo por cada tipo de app.core.field_types (52 tipos)

            reimported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xf-project"},
                files={"upload": ("plantilla_maestra.xlsx", response.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert reimported.status_code == 200, reimported.text
            assert reimported.json()["warnings"] == []
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)
