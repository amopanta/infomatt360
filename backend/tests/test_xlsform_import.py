import json
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.testclient import TestClient

from app.core.security import hash_password
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.assignment import UserProjectAssignment
from app.models.builder import BuilderComponent, BuilderTemplate
from app.models.identity import Project, Role, User


def _build_xlsform(survey_rows: list[list[object]], choices_rows: list[list[object]], survey_headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(survey_headers or ["type", "name", "label"])
    for row in survey_rows:
        survey.append(row)
    choices = workbook.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    for row in choices_rows:
        choices.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def setup_client():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    sessions = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)
    with sessions() as db:
        project = Project(id="xlsform-project", name="XLSForm Project")
        builder_role = Role(id="xlsform-builder-role", name="Builder", permissions="builder.write")
        basic_role = Role(id="xlsform-basic-role", name="Basico", permissions="records.read")
        builder = User(id="xlsform-builder", full_name="Builder", document_id="xlsform-builder-doc", email="xlsform-builder@example.com", password_hash=hash_password("Builder12345!"))
        basic = User(id="xlsform-basic", full_name="Basic", document_id="xlsform-basic-doc", email="xlsform-basic@example.com", password_hash=hash_password("Basic12345!"))
        db.add_all([
            project,
            builder_role,
            basic_role,
            builder,
            basic,
            UserProjectAssignment(user_id=builder.id, project_id=project.id, role_id=builder_role.id, status="active"),
            UserProjectAssignment(user_id=basic.id, project_id=project.id, role_id=basic_role.id, status="active"),
        ])
        db.commit()

    def override_db():
        with sessions() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    return engine, sessions


def auth(client: TestClient, email: str, password: str) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_xlsform_import_maps_common_types_groups_and_repeats():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xlsform-builder@example.com", "Builder12345!")
            basic_headers = auth(client, "xlsform-basic@example.com", "Basic12345!")

            content = _build_xlsform(
                survey_rows=[
                    ["text", "nombre_completo", "Nombre completo"],
                    ["integer", "edad", "Edad"],
                    ["select_one genero", "genero", "Genero"],
                    ["select_multiple ayudas", "ayudas_recibidas", "Ayudas recibidas"],
                    ["geopoint", "ubicacion", "Ubicacion"],
                    ["note", "nota_intro", "Este es solo un texto informativo"],
                    ["begin_group", "grupo_contacto", "Datos de contacto"],
                    ["text", "telefono", "Telefono"],
                    ["end_group", "grupo_contacto", ""],
                    ["begin_repeat", "integrantes_hogar", "Integrantes del hogar"],
                    ["text", "nombre_integrante", "Nombre"],
                    ["integer", "edad_integrante", "Edad"],
                    ["end_repeat", "integrantes_hogar", ""],
                    ["some-exotic-odk-type", "campo_raro", "Campo sin equivalente"],
                ],
                choices_rows=[
                    ["genero", "m", "Masculino"],
                    ["genero", "f", "Femenino"],
                    ["ayudas", "kit", "Kit de herramientas"],
                    ["ayudas", "alimentos", "Alimentos"],
                ],
            )

            denied = client.post(
                "/api/v1/xlsform/import",
                headers=basic_headers,
                data={"project_id": "xlsform-project"},
                files={"upload": ("censo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert denied.status_code == 403

            imported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xlsform-project"},
                files={"upload": ("censo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert imported.status_code == 200
            body = imported.json()
            template_id = body["template_id"]

            # text, integer, select_one, select_multiple, geopoint, telefono (grupo aplanado), repeat, campo_raro = 8
            assert body["imported_fields"] == 8
            assert any("campo_raro" in w and "campo oculto" in w for w in body["warnings"])

            with sessions() as db:
                template = db.query(BuilderTemplate).filter(BuilderTemplate.id == template_id).one()
                assert template.name == "censo"
                assert template.project_id == "xlsform-project"

                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == template_id).order_by(BuilderComponent.sort_order).all()
                by_name = {c.name: c for c in components}

                assert by_name["nombre_completo"].component_type == "TEXT"
                assert by_name["edad"].component_type in ("INTEGER", "NUMBER")
                assert by_name["ubicacion"].component_type == "GPS"
                assert by_name["telefono"].component_type == "TEXT"
                assert "nota_intro" not in by_name
                assert "grupo_contacto" not in by_name

                genero = by_name["genero"]
                assert genero.component_type == "SELECT"
                genero_options = json.loads(genero.config_json)["options"]
                assert {"value": "m", "label": "Masculino"} in genero_options
                assert {"value": "f", "label": "Femenino"} in genero_options

                ayudas = by_name["ayudas_recibidas"]
                assert ayudas.component_type == "MULTISELECT"
                assert len(json.loads(ayudas.config_json)["options"]) == 2

                repeat = by_name["integrantes_hogar"]
                assert repeat.component_type == "REPEAT"
                nested_fields = json.loads(repeat.config_json)["fields"]
                assert [f["name"] for f in nested_fields] == ["nombre_integrante", "edad_integrante"]

                campo_raro = by_name["campo_raro"]
                assert campo_raro.component_type == "HIDDEN"
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_xlsform_import_reads_hint_required_relevant_constraint_range_and_rank_columns():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xlsform-builder@example.com", "Builder12345!")

            headers = ["type", "name", "label", "hint", "required", "relevant", "constraint", "constraint_message", "appearance", "parameters"]
            content = _build_xlsform(
                survey_headers=headers,
                survey_rows=[
                    ["text", "nombre", "Nombre", "Ejemplo: Juan Perez", "yes", "", "", "", "", ""],
                    ["integer", "edad", "Edad", "", "", "", ". >= 0 and . <= 120", "Debe estar entre 0 y 120", "", ""],
                    ["text", "comentario", "Comentario", "", "", "${nombre} != ''", "", "", "", ""],
                    ["range", "satisfaccion", "Satisfaccion", "", "", "", "", "", "", "start=0 end=10 step=1"],
                    ["rank lista_prioridades", "prioridades", "Prioridades", "", "", "", "", "", "", ""],
                    ["background-audio", "audio_fondo", "Audio de fondo", "", "", "", "", "", "", ""],
                ],
                choices_rows=[
                    ["lista_prioridades", "salud", "Salud"],
                    ["lista_prioridades", "educacion", "Educacion"],
                ],
            )

            imported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xlsform-project"},
                files={"upload": ("rico.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert imported.status_code == 200, imported.text
            body = imported.json()
            assert body["imported_fields"] == 5  # nombre, edad, comentario, satisfaccion, prioridades (audio_fondo se omite)
            assert body["warnings"] == []

            with sessions() as db:
                components = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).all()
                by_name = {component.name: component for component in components}

                assert "audio_fondo" not in by_name

                nombre_config = json.loads(by_name["nombre"].config_json)
                assert nombre_config["placeholder"] == "Ejemplo: Juan Perez"
                assert nombre_config["required"] is True

                edad_config = json.loads(by_name["edad"].config_json)
                assert edad_config["min"] == 0.0
                assert edad_config["max"] == 120.0
                assert edad_config["constraint_message"] == "Debe estar entre 0 y 120"

                comentario_config = json.loads(by_name["comentario"].config_json)
                assert comentario_config["relevant"] == {"field": "nombre", "operator": "not_empty", "value": ""}

                assert by_name["satisfaccion"].component_type == "RANGE"
                satisfaccion_config = json.loads(by_name["satisfaccion"].config_json)
                assert satisfaccion_config["min"] == 0.0
                assert satisfaccion_config["max"] == 10.0
                assert satisfaccion_config["step"] == 1.0

                assert by_name["prioridades"].component_type == "RANKING"
                prioridades_config = json.loads(by_name["prioridades"].config_json)
                assert {option["value"] for option in prioridades_config["options"]} == {"salud", "educacion"}
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_xlsform_import_preserves_unrecognized_relevant_and_constraint_expressions_with_warning():
    engine, sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xlsform-builder@example.com", "Builder12345!")

            headers = ["type", "name", "label", "hint", "required", "relevant", "constraint"]
            content = _build_xlsform(
                survey_headers=headers,
                survey_rows=[
                    ["text", "campo_complejo", "Campo", "", "", "${a} > 5 and ${b} < 10", "selected(${lista}, 'x')"],
                ],
                choices_rows=[],
            )

            imported = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xlsform-project"},
                files={"upload": ("complejo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert imported.status_code == 200, imported.text
            body = imported.json()
            assert len(body["warnings"]) == 2
            assert any("relevant" in w for w in body["warnings"])
            assert any("constraint" in w for w in body["warnings"])

            with sessions() as db:
                component = db.query(BuilderComponent).filter(BuilderComponent.template_id == body["template_id"]).one()
                config = json.loads(component.config_json)
                # El texto original se preserva aunque no se haya podido traducir a una condicion/validacion nativa.
                assert config["relevant_expression"] == "${a} > 5 and ${b} < 10"
                assert config["constraint_expression"] == "selected(${lista}, 'x')"
                assert "relevant" not in config
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def test_xlsform_import_rejects_file_without_survey_sheet():
    engine, _sessions = setup_client()
    try:
        with TestClient(app) as client:
            builder_headers = auth(client, "xlsform-builder@example.com", "Builder12345!")
            workbook = Workbook()
            workbook.active.title = "otra_hoja"
            buffer = BytesIO()
            workbook.save(buffer)

            response = client.post(
                "/api/v1/xlsform/import",
                headers=builder_headers,
                data={"project_id": "xlsform-project"},
                files={"upload": ("vacio.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 422
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)
